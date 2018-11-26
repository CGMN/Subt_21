# -*- coding: utf-8 -*-
#python36
import xlrd
import csv
import time
import pandas as pd
from pandas.api.types import CategoricalDtype
import numpy as np
import tkinter.filedialog, re
import openpyxl
from openpyxl.styles import Alignment
#import sys  #esta y las 2 lineas siguientes permiten que trabaje con caracteres especiales
#reload(sys)
#sys.setdefaultencoding('utf-8')

import traceback

f= open("Archivo_de_error.txt","w")

var=0

try:
	#Permite seleccionar el archivo, abre el explorador y guarda la seleccion en la variable file_path
	#______________________________________________________________________________
	root = tkinter.Tk()
	root.withdraw()
	file_path = tkinter.filedialog.askopenfilename()

	#Definiciones___________________________________________________________________
	filtros=["Asig. Urgencia ","Asig. Urgencia (incremento)","Ley 19.536", "Horas extraordinarias",
		"Asignación de turno", "Bonificación compensatoria", "Viáticos", "Función crítica",
		"Asignación de responsabilidad", "Asignación de estímulo", "Experiencia Calificada",
		"Dedicación Exclusiva", "Suplencias y reemplazos"]

#•	Subtítulo 21
#•	Las tres de honorarios
#•	Consultores de llamado
#•	33.000 horas
#•	Ley 19.264
#•	Incremento (Ley 19.264)
#•	Ley 19.536
#•	Horas Extras
#•	Turnos
#•	B. Compensatoria
#•	Viáticos
#•	Función Crítica
#•	A. de Responsabilidad
#•	A. de Estímulos
#•	Experiencia Calificada
#•	Dedicación Exclusiva
#•	Suplencias y Reemplazos


	print ("cantidad de tablas solicitadas: " +str(len(filtros)))

	experimentales=["50 Hospital Padre Alberto Hurtado","51 Centro de Referencia de Salud Maipu",
		"52 Centro de Referencia de Salud Penalolen Cordillera Oriente"]


	pos_tablas=[]
	espacio_de_tablas=40

	for i in range(0,len(filtros)+6):
		if i !=0 and i !=1 and i !=2 and i !=3 and i!=4 and i!=5:
			t=i*espacio_de_tablas
			pos_tablas.append(t)

	#Creacion de tablas_______________________________________________________________

	print ("leyendo base")

	starttime =  time.time()

	df = pd.read_csv(str(file_path), encoding="latin1", low_memory = False)

	#df=df_sin_ordenar.sort_values("mes", ascending=False)

	#Para ordenar los meses al mostrar la tabla
	df['mes'] = df['mes'].astype('category', categories=["enero", "febrero", "marzo", "abril",
		"mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"])

	#dtype = CategoricalDtype(categories=["enero", "febrero", "marzo", "abril", "mayo","junio","julio",
	#	"agosto","septiembre","octubre","noviembre","diciembre"], ordered=True)

	#df.astype(dtype)


	print (time.time()-starttime)

	print ("base leida")
	print (len(df))
	#Para tener una lista con las cabeceras, _____________________________________

	#Para tener un listado con las clasificaciones y despues usarlas para que no muera el programa si falta alguna
	clasificaciones_todas=[]
	for i in range(0, len(df)):
		clasificaciones_todas.append(df.CLASIFICACION[i])

	clasificaciones=list(set(clasificaciones_todas))
	#print len(clasificaciones)

	#Quitamos de filtros lo que no está en el archivo______________________________
	quitar_de_filtros=[]
	for i in range(0,len(filtros)):
		if filtros[i] not in clasificaciones:
			quitar_de_filtros.append(filtros[i])
			#print len(filtros)

	for i in range(0,len(quitar_de_filtros)):
		a=filtros.index(quitar_de_filtros[i])
		filtros.insert(a,"0")
		filtros.remove(quitar_de_filtros[i])

	print ("cantidad de tablas a crear: "+ str(len(filtros)))

	if len(quitar_de_filtros)>0:
		print ("En el archivo sigfe no se encuentra: ")
		for i in range(0,len(quitar_de_filtros)):
			print (" ",quitar_de_filtros[i])

	print ("creando tablas")
	#creacion de tablas especificas, las 4 primeras______________________________________________________________
	tablas=[]

	tabla4=pd.pivot_table(df[df.SubTitulo=="21 GASTOS EN PERSONAL"], index=["Institucion"],
			columns=["mes"],	values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')

	if "Honorarios asim. Ley 18.834" in clasificaciones:
		table1=pd.pivot_table(df[df.CLASIFICACION=="Honorarios asim. Ley 18.834"], index=["Institucion"],
			columns=["mes"],	values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')

	if "Honorarios asim. Ley Médica" in clasificaciones:
		table2=pd.pivot_table(df[df.CLASIFICACION=="Honorarios asim. Ley Médica"], index=["Institucion"],
			columns=["mes"],	values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')

	if "Honorarios asim. Ley Médica" in clasificaciones and "Honorarios asim. Ley 18.834" in clasificaciones:
		table3=pd.pivot_table(df[(df.CLASIFICACION=="Honorarios asim. Ley Médica") | (df.CLASIFICACION=="Honorarios asim. Ley 18.834")],
			index=["Institucion"], columns=["mes"],	values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')

	#consultores de llamada
	table5=pd.pivot_table(df[(df.SubTitulo=="21 GASTOS EN PERSONAL")&(df.Item=="03 Otras Remuneraciones")&(df.Asignación=="001 Honorarios a Suma Alzada - Personas Naturales")
		&(df.SubAsignación=="001 Honorarios A Suma Alzada Personas Naturales")&(df.Específico=="01 Hon Conv Tratantes O Consult Llamadas Art 24 L 19664")],
			index=["Institucion"], columns=["mes"],values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')

	#33 mil horas
	table6=pd.pivot_table(df[(df.SubTitulo=="21 GASTOS EN PERSONAL")&(df.Item=="03 Otras Remuneraciones")&(df.Asignación=="001 Honorarios a Suma Alzada - Personas Naturales")
		&(df.SubAsignación=="001 Honorarios A Suma Alzada Personas Naturales")&(df.Específico=="06 Personal Médico Programa Cierre De Brechas")],
			index=["Institucion"], columns=["mes"],values=["Devengado"],aggfunc=[np.sum], fill_value=0, margins=True, margins_name='Total')


	#creacion de tablas generales_________________________________________________________________________________
	for i in range(0,len(filtros)):#index=filas

		table = pd.pivot_table(df[df.CLASIFICACION==filtros[i]],index=["Institucion"], columns=["mes"],
			values=["Devengado"],aggfunc=[np.sum],fill_value=0,margins=True,margins_name='Total')
		tablas.append(table)

	#Crea el archivo excel________________________________________________________
	writer = pd.ExcelWriter('Resumen_glosas.xlsx')

	#para posicionar las tablas en el archivo______________________________________

	tabla4.to_excel(writer,"Hoja1", startcol=1, startrow=0)

	if "Honorarios asim. Ley 18.834" in clasificaciones:
		table1.to_excel(writer,"Hoja1", startcol=1, startrow=40)

	if "Honorarios asim. Ley Médica" in clasificaciones:
		table2.to_excel(writer,"Hoja1", startcol=1, startrow=80)

	if "Honorarios asim. Ley Médica" in clasificaciones and "Honorarios asim. Ley 18.834" in clasificaciones:
		table3.to_excel(writer,"Hoja1", startcol=1, startrow=120)

	table5.to_excel(writer,"Hoja1",startcol=1, startrow=160)

	table6.to_excel(writer,"Hoja1",startcol=1,startrow=200)

	for i in range(0,len(tablas)):
		tablas[i].to_excel(writer,"Hoja1", startcol=1, startrow=pos_tablas[i])

	#Graba el archivo_____________________________________________________________
	writer.save()


	print ("Resumen_glosas creado")

	#Para darle formato al archivo________________________________________________________
	doc = openpyxl.load_workbook('Resumen_glosas.xlsx')

	H=doc.get_sheet_names() #la variable H guarda un listado con las hojas del excel
	hoja1 = doc.get_sheet_by_name(str(H[0])) #con H[0], le decimos que use la primera hoja del excel

	#Ancho de columna___________________________________________________________________________
	hoja1.column_dimensions['B'].width = 60

	#hoja1.unmerge_cells("B1:C1")
	#hoja1.merge_cells('B1:C1')# me falta

	#Para poner el codigo del servicio al lado"_______________________________________________
	excluir=["Total", "mes", "Institucion", filtros]

	for i in range(5,pos_tablas[-1]+espacio_de_tablas):
		if hoja1.cell(row=i,column=2).value!=None:
			if hoja1.cell(row=i,column=2).value not in excluir:
				hoja1["A"+str(i)]=int(hoja1.cell(row=i,column=2).value[0:2])

	#alineacion a la izquierda______________________________________________________________
	for row in hoja1['B1':'B'+str(pos_tablas[-1]+espacio_de_tablas)]:
		for cell in row:
		 	cell.alignment = Alignment(horizontal="left")

	#cabeceras__________________________________________________________

	hoja1["B1"]="21 GASTOS EN PERSONAL"

	if "Honorarios asim. Ley 18.834" in clasificaciones:
		hoja1["B41"]="Honorarios asim. Ley 18.834 sin experimentales"

	if "Honorarios asim. Ley Médica" in clasificaciones:
		hoja1["B81"]="Honorarios asim. Ley Médica sin experimentales"

	if "Honorarios asim. Ley Médica" in clasificaciones and "Honorarios asim. Ley 18.834" in clasificaciones:
		hoja1["B121"]="Honorarios Ley 18.834 y Ley 19.664 solo experimentales"

	hoja1["B161"]="Consultores de Llamada"

	hoja1["B201"]="33.000 horas"

	for i in range(0,len(filtros)):
		hoja1["B"+str(pos_tablas[i]+1)]=filtros[i]


	#Para borrar los datos de las tablas en 0
	for i in range(1,pos_tablas[-1]+espacio_de_tablas):
		if hoja1["B"+str(i)].value==0:
			for j in hoja1["C"+str(i+4):"O"+str(i+38)]:
				for k in j:
					k.value=""

	#Para borrar los totales de las tablas ubicadas en segunda, tercera y cuarta ubicacion
	for i in range(43,158):
		if hoja1["B"+str(i)].value=="Total":
			for j in hoja1["C"+str(i):"O"+str(i)]:
				for k in j:
					k.value=""

	#Para borrar los experimentales de la tablas ubicadas en segunda y tercera posicion del archivo
	for i in range(43,118):
		if hoja1["A"+str(i)].value==50 or hoja1["A"+str(i)].value==51 or hoja1["A"+str(i)].value==52:
			for j in hoja1["C"+str(i):"O"+str(i)]:
				for k in j:
					k.value=""

	#borrar los otros servicios y dejando los experimentales
	todos_los_servicios=[20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,
		38,39,40,41,42,43,44,45,46,47,53]

	for i in range(125,158):
		for k in range(0,len(todos_los_servicios)):
			if hoja1["A"+str(i)].value==todos_los_servicios[k]:
				for j in hoja1["C"+str(i):"O"+str(i)]:
					for k in j:
						k.value=""

	#graba archivo
	doc.save("Resumen_glosas_SS.xlsx")

	print ("Resumen_glosas modificado")

except:
	var = traceback.format_exc()
	print (var)
if var !=0:
	f.write(var)
	f.close()
	print ("")
	print (" X X X X X X X X X X X X X X X X X X ")
	print ("")
	print ("El programa arroja un error")
	print ("Se ha creado un archivo llamado 'Archivo_de_error' con el detalle")
	print ("")
	print (" X X X X X X X X X X X X X X X X X X ")

f.close()
print ("")
print ("proceso terminado")

time.sleep(6)
